"""
Excel Handler Service
Handles reading and writing Excel files with Google Drive integration
"""

import pandas as pd
import openpyxl
import tempfile
import os
import logging
from io import BytesIO

logger = logging.getLogger(__name__)

def read_excel_from_drive(file_id, sheet_name=None):
    """
    Read Excel file from Google Drive into pandas DataFrame

    Args:
        file_id: Google Drive file ID (string)
        sheet_name: Optional sheet name (defaults to first sheet)

    Returns:
        pandas.DataFrame: Excel data
    """
    try:
        # Import here to avoid circular imports
        from services import google_drive

        # Ensure file_id is a clean string
        if isinstance(file_id, bytes):
            file_id = file_id.decode('utf-8')
        file_id = str(file_id).strip()

        logger.info(f"Reading Excel file with ID: {file_id}")

        # Download file as bytes
        file_bytes = google_drive.download_file_as_bytes(file_id)

        # Read with pandas from bytes
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name or 0)

        logger.info(f"Excel file read successfully: {len(df)} rows from {file_id}")
        return df

    except Exception as e:
        logger.error(f"Error reading Excel from Drive: {e}")
        raise

def write_excel_to_drive(df, file_id, sheet_name='Sheet1'):
    """
    Write pandas DataFrame to Excel file on Google Drive

    Args:
        df: pandas DataFrame
        file_id: Google Drive file ID to update
        sheet_name: Sheet name to write to

    Returns:
        dict: Updated file metadata
    """
    try:
        from services import google_drive

        # Ensure file_id is a clean string
        if isinstance(file_id, bytes):
            file_id = file_id.decode('utf-8')
        file_id = str(file_id).strip()

        # Write to temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        df.to_excel(tmp_path, sheet_name=sheet_name, index=False)

        # Upload to Drive
        result = google_drive.upload_file(tmp_path, file_id=file_id)

        # Clean up
        os.remove(tmp_path)

        logger.info(f"Wrote {len(df)} rows to file {file_id}")
        return result

    except Exception as e:
        logger.error(f"Error writing Excel to Drive: {e}")
        raise

def update_cell_in_drive(file_id, sheet_name, cell, value):
    """
    Update a single cell in Excel file on Google Drive

    Args:
        file_id: Google Drive file ID (string)
        sheet_name: Sheet name (e.g., 'Data')
        cell: Cell reference (e.g., 'B2')
        value: New value for the cell

    Returns:
        dict: Updated file metadata
    """
    try:
        from services import google_drive

        # Ensure file_id is a clean string
        if isinstance(file_id, bytes):
            file_id = file_id.decode('utf-8')
        file_id = str(file_id).strip()

        logger.info(f"Updating cell {cell} in file {file_id}")

        # Download file to temp location
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        google_drive.download_file(file_id, tmp_path)

        # Update cell using openpyxl
        wb = openpyxl.load_workbook(tmp_path)

        if sheet_name not in wb.sheetnames:
            available = wb.sheetnames
            wb.close()
            os.remove(tmp_path)
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {available}"
            )

        ws = wb[sheet_name]
        ws[cell] = value
        wb.save(tmp_path)
        wb.close()

        logger.info(f"Updated cell {cell} to '{value}' in sheet '{sheet_name}'")

        # Upload back to Drive
        result = google_drive.upload_file(tmp_path, file_id=file_id)

        # Clean up
        os.remove(tmp_path)

        logger.info(f"Cell {cell} updated successfully in {file_id}")
        return result

    except Exception as e:
        logger.error(f"Error updating cell in Drive: {e}")
        raise

def append_rows_to_drive(file_id, sheet_name, new_data):
    """
    Append rows to existing Excel file on Google Drive
    """
    try:
        # Ensure file_id is a clean string
        if isinstance(file_id, bytes):
            file_id = file_id.decode('utf-8')
        file_id = str(file_id).strip()

        existing_df = read_excel_from_drive(file_id, sheet_name)
        combined_df = pd.concat([existing_df, new_data], ignore_index=True)
        result = write_excel_to_drive(combined_df, file_id, sheet_name)

        logger.info(f"Appended {len(new_data)} rows to file {file_id}")
        return result

    except Exception as e:
        logger.error(f"Error appending rows: {e}")
        raise

def get_excel_info(file_id):
    """
    Get information about Excel file (sheets, dimensions)
    """
    try:
        from services import google_drive

        # Ensure file_id is a clean string
        if isinstance(file_id, bytes):
            file_id = file_id.decode('utf-8')
        file_id = str(file_id).strip()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        google_drive.download_file(file_id, tmp_path)

        wb = openpyxl.load_workbook(tmp_path, read_only=True)

        sheets_info = {}
        for name in wb.sheetnames:
            ws = wb[name]
            sheets_info[name] = {
                'max_row': ws.max_row,
                'max_column': ws.max_column
            }

        wb.close()
        os.remove(tmp_path)

        return {
            'file_id': file_id,
            'sheets': sheets_info,
            'sheet_count': len(sheets_info)
        }

    except Exception as e:
        logger.error(f"Error getting Excel info: {e}")
        raise
