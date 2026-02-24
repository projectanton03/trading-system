"""
Trading System API - Full System
Flask application with Google Drive integration and complete data pipeline

Week 1: Foundation setup with template management
"""

from flask import Flask, request, jsonify
import os
import sys
from datetime import datetime
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Import configurations
try:
    import config
    config.validate_config()
    logger.info("Configuration validated successfully")
except Exception as e:
    logger.error(f"Configuration error: {e}")
    # Continue but warn

# Import Google Drive service
try:
    from services import google_drive
    logger.info("Google Drive service loaded")
except ImportError as e:
    logger.error(f"Failed to import google_drive: {e}")
    google_drive = None

# Import Excel handler (optional for now)
try:
    from services import excel_handler
    logger.info("Excel handler loaded")
except ImportError as e:
    logger.warning(f"Excel handler not available: {e}")
    excel_handler = None

#═══════════════════════════════════════════════════════════════════════════════
# HEALTH CHECK ENDPOINTS
#═══════════════════════════════════════════════════════════════════════════════

@app.route('/')
def home():
    """Main health check endpoint"""
    return jsonify({
        'status': 'running',
        'message': 'Trading System API - Full System',
        'version': '1.0.0',
        'timestamp': datetime.now().isoformat(),
        'week': 1,
        'phase': 'Infrastructure Setup'
    })

@app.route('/health')
def health():
    """Detailed health check with all systems"""
    health_status = {
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'components': {}
    }
    
    # Check API keys
    health_status['components']['api_keys'] = {
        'groq': bool(os.environ.get('GROQ_API_KEY')),
        'alpha_vantage': bool(os.environ.get('ALPHA_VANTAGE_KEY')),
        'fred': bool(os.environ.get('FRED_API_KEY')),
        'telegram': bool(os.environ.get('TELEGRAM_BOT_TOKEN'))
    }
    
    # Check Google Drive connection
    if google_drive:
        try:
            success, message = google_drive.test_drive_connection()
            health_status['components']['google_drive'] = {
                'status': 'connected' if success else 'failed',
                'message': message
            }
        except Exception as e:
            health_status['components']['google_drive'] = {
                'status': 'error',
                'message': str(e)
            }
    else:
        health_status['components']['google_drive'] = {
            'status': 'error',
            'message': 'Google Drive module not loaded'
        }

    # Overall status
    all_healthy = all(
        comp.get('status') != 'failed'
        for comp in health_status['components'].values()
        if isinstance(comp, dict) and 'status' in comp
    )
    
    if not all_healthy:
        health_status['status'] = 'degraded'
    
    return jsonify(health_status)


#═══════════════════════════════════════════════════════════════════════════════
# GOOGLE DRIVE ENDPOINTS (Week 1 Focus)
#═══════════════════════════════════════════════════════════════════════════════

@app.route('/drive/list/<folder_type>')
def list_drive_folder(folder_type):
    """
    List files in a Google Drive folder
    """
    try:
        folder_id = config.DRIVE_FOLDERS.get(folder_type)

        if not folder_id:
            return jsonify({
                'error': f'Unknown folder type: {folder_type}',
                'available_types': list(config.DRIVE_FOLDERS.keys())
            }), 400

        files = google_drive.list_files_in_folder(folder_id)

        return jsonify({
            'folder_type': folder_type,
            'folder_id': folder_id,
            'file_count': len(files),
            'files': files,
            'status': 'success'
        })

    except Exception as e:
        logger.error(f"Error listing folder: {e}")
        return jsonify({'error': str(e)}), 500


import numpy as np
import pandas as pd
from io import BytesIO

@app.route('/drive/read/<file_id>')
def read_excel_file(file_id):
    """
    Read Excel file from Google Drive
    """
    try:
        if not google_drive:
            return jsonify({'error': 'Google Drive not available'}), 500

        # Download file as binary bytes
        file_bytes = google_drive.download_file_as_bytes(file_id)

        # Read directly from bytes using BytesIO - NO utf-8 decode!
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=0)

        # Make dataframe JSON safe
        df = df.replace({np.nan: None})
        df = df.where(pd.notnull(df), None)

        # Convert columns to strings (handles special types)
        columns = [str(c) for c in df.columns]

        preview = df.head(10).to_dict('records')

        return jsonify({
            'file_id': file_id,
            'rows': int(len(df)),
            'columns': columns,
            'preview': preview,
            'status': 'success'
        })

    except Exception as e:
        logger.error(f"Error reading file {file_id}: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/templates/test-update', methods=['POST'])
def test_template_update():
    """
    Test updating a template in Google Drive
    """
    try:
        data = request.get_json()

        file_id = data.get('file_id')
        sheet_name = data.get('sheet_name', 'Sheet1')
        cell = data.get('cell', 'A1')
        value = data.get('value')

        if not value:
            value = f'Updated {datetime.now().isoformat()}'

        excel_handler.update_cell_in_drive(file_id, sheet_name, cell, value)

        return jsonify({
            'message': f'Cell {cell} updated successfully',
            'file_id': file_id,
            'sheet_name': sheet_name,
            'new_value': value,
            'status': 'success'
        })

    except Exception as e:
        logger.error(f"Error updating template: {e}")
        return jsonify({'error': str(e)}), 500


#═══════════════════════════════════════════════════════════════════════════════
# MACRO DATA ENDPOINTS (Week 1-2 implementation)
#═══════════════════════════════════════════════════════════════════════════════

@app.route('/macro/templates')
def list_macro_templates():
    """List all macro indicator templates from Google Drive"""
    try:
        templates = []
        
        for folder_type in ['macro_leading', 'macro_coincident', 'macro_international']:
            folder_id = config.DRIVE_FOLDERS.get(folder_type)
            if folder_id:
                files = google_drive.list_files_in_folder(folder_id)
                templates.extend([{
                    **f,
                    'category': folder_type
                } for f in files])
        
        return jsonify({
            'total_templates': len(templates),
            'templates': templates,
            'status': 'success'
        })
        
    except Exception as e:
        logger.error(f"Error listing templates: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/macro/test-fetch', methods=['GET'])
def test_macro_fetch():
    """
    Test endpoint - Returns dummy macro data
    Week 2: Replace with real data fetching
    """
    logger.info("Fetching macro data (test mode - Week 1)")
    
    macro_data = {
        'ism_manufacturing': 54.2,
        'ism_services': 52.8,
        'yield_10y': 4.25,
        'yield_2y': 4.65,
        'yield_spread': -0.40,
        'credit_spread_bbb': 2.80,
        'consumer_confidence': 103.5,
        'building_permits': 1450000,
        'housing_starts': 1420000,
        'copper_price_change': 3.2,
        'china_pmi': 49.8,
        'timestamp': datetime.now().isoformat(),
        'status': 'success',
        'mode': 'test',
        'note': 'Week 1 - Dummy data. Week 2 will implement real fetching.'
    }
    
    return jsonify(macro_data)

@app.route('/macro/fetch', methods=['GET'])
def fetch_macro_data():
    """
    Fetch REAL macro data from FRED API
    Week 2: This replaces the test endpoint with actual data
    """
    try:
        from services.fred_api import FREDClient, calculate_yield_spread, calculate_credit_spread
        
        logger.info("Fetching real macro data from FRED API")
        
        # Initialize FRED client
        client = FREDClient()
        
        # Fetch key indicators
        treasury_10y, date_10y = client.get_latest_value('DGS10')
        treasury_2y, date_2y = client.get_latest_value('DGS2')
        consumer_sentiment, date_conf = client.get_latest_value('UMCSENT')
        building_permits, date_permits = client.get_latest_value('PERMIT')
        housing_starts, date_starts = client.get_latest_value('HOUST')
        baa_corporate, date_baa = client.get_latest_value('DBAA')
        industrial_production, date_indpro = client.get_latest_value('INDPRO')
        
        # Calculate spreads
        yield_spread = calculate_yield_spread(treasury_10y, treasury_2y)
        credit_spread = calculate_credit_spread(baa_corporate, treasury_10y)
        
        # Build response
        macro_data = {
            'treasury_10y': treasury_10y,
            'treasury_2y': treasury_2y,
            'yield_spread': yield_spread,
            'credit_spread_bbb': credit_spread,
            'consumer_sentiment': consumer_sentiment,
            'building_permits': building_permits,
            'housing_starts': housing_starts,
            'baa_corporate': baa_corporate,
            'industrial_production': industrial_production,
            'timestamp': datetime.now().isoformat(),
            'status': 'success',
            'source': 'FRED API',
            'note': 'Week 2 - Real data from Federal Reserve Economic Data',
            'data_dates': {
                'treasury_10y': date_10y,
                'treasury_2y': date_2y,
                'consumer_sentiment': date_conf,
                'building_permits': date_permits,
                'housing_starts': date_starts,
                'baa_corporate': date_baa,
                'industrial_production': date_indpro
            }
        }
        
        logger.info(f"Successfully fetched macro indicators from FRED")
        return jsonify(macro_data)
        
    except Exception as e:
        logger.error(f"Error fetching macro data: {e}")
        return jsonify({
            'error': str(e),
            'status': 'failed',
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/macro/audit-templates', methods=['GET'])
def audit_templates():
    """
    Audit all Excel templates to determine last update dates and backfill requirements
    Week 2: Step 1 before backfilling data
    """
    try:
        logger.info("Starting template audit...")
        
        # Known file IDs that we have
        templates_to_audit = {
            'ISM_Manufacturing': '1o8eHxS_8V-tOgW_4lrOMCZ9FGCclGyrO',
            'US_Sector_Data': '11UwhrI8uUdr7ngWy_87rizWBEejLCdqo',
            'Benchmark_Yields_US': '1I3f36ghjh-NpI_EyhlZ9JTNUnGIWDkg4',
        }
        
        results = []
        
        for template_name, file_id in templates_to_audit.items():
            try:
                # Download file
                file_bytes = google_drive.download_file_as_bytes(file_id)
                
                # Read Excel - try first sheet
                df = pd.read_excel(BytesIO(file_bytes), sheet_name=0)
                
                # Find date column (try common names)
                date_col = None
                for col in df.columns:
                    col_lower = str(col).lower()
                    if any(pattern in col_lower for pattern in ['date', 'period', 'month', 'year', 'time']):
                        date_col = col
                        break
                
                if not date_col:
                    # Assume first column is date
                    date_col = df.columns[0]
                
                # Convert to datetime
                df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
                df_valid = df[df[date_col].notna()]
                
                if len(df_valid) == 0:
                    results.append({
                        'template_name': template_name,
                        'status': 'ERROR',
                        'error': 'No valid dates found'
                    })
                    continue
                
                # Get date range
                start_date = df_valid[date_col].min()
                end_date = df_valid[date_col].max()
                total_rows = len(df_valid)
                
                # Calculate gap from last update to now
                current_date = datetime.now()
                gap_days = (current_date - end_date).days
                
                # Estimate frequency
                date_diffs = df_valid[date_col].diff().dropna()
                if len(date_diffs) > 0:
                    avg_diff_days = date_diffs.dt.days.median()
                    
                    if avg_diff_days <= 1.5:
                        frequency = 'daily'
                        gap_periods = gap_days
                    elif avg_diff_days <= 35:
                        frequency = 'monthly'
                        gap_periods = gap_days // 30
                    elif avg_diff_days <= 100:
                        frequency = 'quarterly'
                        gap_periods = gap_days // 90
                    else:
                        frequency = 'annual'
                        gap_periods = gap_days // 365
                else:
                    frequency = 'unknown'
                    gap_periods = 0
                
                results.append({
                    'template_name': template_name,
                    'file_id': file_id,
                    'status': 'SUCCESS',
                    'start_date': start_date.strftime('%Y-%m-%d'),
                    'end_date': end_date.strftime('%Y-%m-%d'),
                    'last_update_days_ago': gap_days,
                    'total_rows': int(total_rows),
                    'frequency': frequency,
                    'gap_periods': int(gap_periods),
                    'date_column': str(date_col),
                    'needs_backfill': gap_days > 7
                })
                
                logger.info(f"Audited {template_name}: {total_rows} rows, last update {end_date.strftime('%Y-%m-%d')}")
                
            except Exception as e:
                logger.error(f"Error auditing {template_name}: {e}")
                results.append({
                    'template_name': template_name,
                    'status': 'ERROR',
                    'error': str(e)
                })
        
        # Generate summary
        total = len(results)
        success = sum(1 for r in results if r['status'] == 'SUCCESS')
        needs_backfill = sum(1 for r in results if r.get('needs_backfill', False))
        
        total_gap_days = sum(r.get('last_update_days_ago', 0) for r in results if r['status'] == 'SUCCESS')
        avg_gap_days = total_gap_days // success if success > 0 else 0
        
        return jsonify({
            'status': 'success',
            'timestamp': datetime.now().isoformat(),
            'summary': {
                'total_templates_checked': total,
                'successfully_audited': success,
                'needs_backfill': needs_backfill,
                'average_gap_days': avg_gap_days
            },
            'templates': results,
            'next_steps': [
                'Review audit results',
                'Verify file IDs for all 51 templates',
                'Execute backfill for templates with gaps',
                'Set up weekly incremental updates'
            ]
        })
        
    except Exception as e:
        logger.error(f"Error in template audit: {e}")
        return jsonify({
            'error': str(e),
            'status': 'failed'
        }), 500

@app.route('/macro/audit-templates-v2', methods=['GET'])
def audit_templates_v2():
    """
    Enhanced template audit - checks multiple sheets and better date detection
    """
    try:
        import openpyxl
        import tempfile
        
        logger.info("Starting enhanced template audit...")
        
        templates_to_audit = {
            'ISM_Manufacturing': '1o8eHxS_8V-tOgW_4lrOMCZ9FGCclGyrO',
            'US_Sector_Data': '11UwhrI8uUdr7ngWy_87rizWBEejLCdqo',
            'Benchmark_Yields_US': '1I3f36ghjh-NpI_EyhlZ9JTNUnGIWDkg4',
        }
        
        results = []
        
        for template_name, file_id in templates_to_audit.items():
            try:
                # Download file to temp location
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    tmp_path = tmp.name
                
                google_drive.download_file(file_id, tmp_path)
                
                # Load workbook to see all sheets
                wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                sheet_names = wb.sheetnames
                
                logger.info(f"{template_name} has sheets: {sheet_names}")
                
                # Try to find a sheet with data (not 'NOTES' or 'README')
                data_sheets = [s for s in sheet_names if s.upper() not in ['NOTES', 'README', 'INFO', 'INSTRUCTIONS']]
                
                if not data_sheets:
                    data_sheets = sheet_names
                
                # Try each sheet until we find dates
                best_result = None
                
                for sheet_name in data_sheets[:3]:  # Try first 3 data sheets
                    try:
                        df = pd.read_excel(tmp_path, sheet_name=sheet_name)
                        
                        # Skip if too few rows
                        if len(df) < 5:
                            continue
                        
                        # Try to find date column more intelligently
                        date_col = None
                        
                        # Method 1: Look for columns with 'date' in name
                        for col in df.columns:
                            col_str = str(col).lower()
                            if any(word in col_str for word in ['date', 'period', 'month', 'year', 'time', 'day']):
                                # Try to parse this column
                                test_series = pd.to_datetime(df[col], errors='coerce')
                                valid_dates = test_series.notna().sum()
                                if valid_dates > len(df) * 0.5:  # At least 50% valid dates
                                    date_col = col
                                    break
                        
                        # Method 2: Try first few columns
                        if not date_col:
                            for col in df.columns[:3]:
                                test_series = pd.to_datetime(df[col], errors='coerce')
                                valid_dates = test_series.notna().sum()
                                if valid_dates > len(df) * 0.5:
                                    date_col = col
                                    break
                        
                        if not date_col:
                            continue
                        
                        # Parse dates
                        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
                        df_valid = df[df[date_col].notna()]
                        
                        if len(df_valid) < 5:
                            continue
                        
                        # Get date range
                        start_date = df_valid[date_col].min()
                        end_date = df_valid[date_col].max()
                        
                        # Skip if dates are clearly wrong (before 1900 or after 2030)
                        if start_date.year < 1900 or end_date.year > 2030:
                            continue
                        
                        total_rows = len(df_valid)
                        current_date = datetime.now()
                        gap_days = (current_date - end_date).days
                        
                        # Calculate frequency
                        date_diffs = df_valid[date_col].diff().dropna()
                        if len(date_diffs) > 0:
                            avg_diff_days = date_diffs.dt.days.median()
                            
                            if avg_diff_days <= 1.5:
                                frequency = 'daily'
                                gap_periods = gap_days
                            elif avg_diff_days <= 35:
                                frequency = 'monthly'
                                gap_periods = gap_days // 30
                            elif avg_diff_days <= 100:
                                frequency = 'quarterly'
                                gap_periods = gap_days // 90
                            else:
                                frequency = 'annual'
                                gap_periods = gap_days // 365
                        else:
                            frequency = 'unknown'
                            gap_periods = 0
                        
                        result = {
                            'template_name': template_name,
                            'file_id': file_id,
                            'status': 'SUCCESS',
                            'sheet_name': sheet_name,
                            'start_date': start_date.strftime('%Y-%m-%d'),
                            'end_date': end_date.strftime('%Y-%m-%d'),
                            'last_update_days_ago': gap_days,
                            'total_rows': int(total_rows),
                            'frequency': frequency,
                            'gap_periods': int(gap_periods),
                            'date_column': str(date_col),
                            'needs_backfill': gap_days > 7,
                            'all_sheets': sheet_names
                        }
                        
                        # Keep best result (most rows, most recent end date)
                        if best_result is None or total_rows > best_result['total_rows']:
                            best_result = result
                        
                    except Exception as e:
                        logger.warning(f"Could not parse sheet {sheet_name}: {e}")
                        continue
                
                wb.close()
                os.remove(tmp_path)
                
                if best_result:
                    results.append(best_result)
                    logger.info(f"Audited {template_name}: {best_result['total_rows']} rows, ends {best_result['end_date']}")
                else:
                    results.append({
                        'template_name': template_name,
                        'status': 'ERROR',
                        'error': 'No valid date columns found in any sheet',
                        'sheets_checked': sheet_names
                    })
                
            except Exception as e:
                logger.error(f"Error auditing {template_name}: {e}")
                results.append({
                    'template_name': template_name,
                    'status': 'ERROR',
                    'error': str(e)
                })
        
        # Generate summary
        total = len(results)
        success = sum(1 for r in results if r['status'] == 'SUCCESS')
        needs_backfill = sum(1 for r in results if r.get('needs_backfill', False))
        
        if success > 0:
            total_gap_days = sum(r.get('last_update_days_ago', 0) for r in results if r['status'] == 'SUCCESS')
            avg_gap_days = total_gap_days // success
        else:
            avg_gap_days = 0
        
        return jsonify({
            'status': 'success',
            'timestamp': datetime.now().isoformat(),
            'summary': {
                'total_templates_checked': total,
                'successfully_audited': success,
                'needs_backfill': needs_backfill,
                'average_gap_days': avg_gap_days,
                'oldest_end_date': min([r['end_date'] for r in results if r['status'] == 'SUCCESS']) if success > 0 else None,
                'newest_end_date': max([r['end_date'] for r in results if r['status'] == 'SUCCESS']) if success > 0 else None
            },
            'templates': results
        })
        
    except Exception as e:
        logger.error(f"Error in template audit: {e}")
        return jsonify({
            'error': str(e),
            'status': 'failed'
        }), 500

@app.route('/macro/backfill-yields', methods=['POST'])
def backfill_yields_endpoint():
    """
    Backfill Benchmark Yields US with FRED data
    Week 2: Fill 5-year gap (Jan 2021 → Feb 2026)
    """
    try:
        import openpyxl
        import tempfile
        from services.fred_api import FREDClient
        
        logger.info("Starting Benchmark Yields backfill...")
        
        file_id = '1I3f36ghjh-NpI_EyhlZ9JTNUnGIWDkg4'
        start_date = '2021-01-04'
        end_date = datetime.now().strftime('%Y-%m-%d')
        
        # Mapping of sheet names to FRED series IDs
        yield_series = {
            '3mo': 'DGS3MO',
            '6mo': 'DGS6MO',
            '1yr': 'DGS1',
            '2yr': 'DGS2',
            '3yr': 'DGS3',
            '5yr': 'DGS5',
            '7yr': 'DGS7',
            '10yr': 'DGS10',
            '20yr': 'DGS20',
            '30yr': 'DGS30',
        }
        
        # Initialize FRED client
        client = FREDClient()
        
        # Download file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        google_drive.download_file(file_id, tmp_path)
        
        # Load workbook
        wb = openpyxl.load_workbook(tmp_path)
        
        results = []
        total_rows_added = 0
        
        # Process each yield series
        for sheet_name, series_id in yield_series.items():
            try:
                if sheet_name not in wb.sheetnames:
                    logger.warning(f"Sheet {sheet_name} not found")
                    continue
                
                # Fetch data from FRED
                data = client.get_series(series_id, limit=2000, sort_order='asc')
                
                if not data or 'observations' not in data:
                    continue
                
                # Convert to DataFrame
                df = pd.DataFrame(data['observations'])
                df['date'] = pd.to_datetime(df['date'])
                df['value'] = pd.to_numeric(df['value'], errors='coerce')
                
                # Filter to backfill range
                df = df[df['date'] >= start_date]
                df = df[df['date'] <= end_date]
                df = df.dropna(subset=['value'])
                
                if len(df) == 0:
                    continue
                
                # Read existing sheet
                ws = wb[sheet_name]
                
                # Find last row with data
                last_row = ws.max_row
                
                # Append new data
                rows_added = 0
                for _, row in df.iterrows():
                    last_row += 1
                    ws.cell(row=last_row, column=1, value=row['date'])
                    ws.cell(row=last_row, column=2, value=row['value'])
                    rows_added += 1
                
                total_rows_added += rows_added
                
                results.append({
                    'sheet': sheet_name,
                    'series_id': series_id,
                    'rows_added': rows_added,
                    'date_range': f"{df['date'].min().strftime('%Y-%m-%d')} to {df['date'].max().strftime('%Y-%m-%d')}"
                })
                
                logger.info(f"Added {rows_added} rows to {sheet_name}")
                
            except Exception as e:
                logger.error(f"Error processing {sheet_name}: {e}")
        
        # Save workbook
        wb.save(tmp_path)
        wb.close()
        
        # Upload back to Drive
        google_drive.upload_file(tmp_path, file_id=file_id)
        
        # Clean up
        os.remove(tmp_path)
        
        return jsonify({
            'status': 'success',
            'message': f'Backfilled {total_rows_added} rows across {len(results)} sheets',
            'file_id': file_id,
            'start_date': start_date,
            'end_date': end_date,
            'sheets_updated': results,
            'total_rows_added': total_rows_added,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error in yields backfill: {e}")
        return jsonify({
            'error': str(e),
            'status': 'failed'
        }), 500

@app.route('/macro/inspect-yields', methods=['GET'])
def inspect_yields():
    """
    Inspect the Benchmark Yields file structure to understand layout
    """
    try:
        import openpyxl
        import tempfile
        
        logger.info("Inspecting Benchmark Yields file structure...")
        
        file_id = '1I3f36ghjh-NpI_EyhlZ9JTNUnGIWDkg4'
        
        # Download file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        google_drive.download_file(file_id, tmp_path)
        
        # Load workbook
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        
        inspection = {
            'file_id': file_id,
            'total_sheets': len(wb.sheetnames),
            'all_sheets': wb.sheetnames,
            'sheet_details': []
        }
        
        # Inspect first few sheets
        sheets_to_check = ['10yr', '2yr', '5yr', '3mo']
        
        for sheet_name in sheets_to_check:
            if sheet_name not in wb.sheetnames:
                inspection['sheet_details'].append({
                    'sheet': sheet_name,
                    'status': 'NOT_FOUND'
                })
                continue
            
            ws = wb[sheet_name]
            
            # Get sheet dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            
            # Sample first 10 rows
            sample_rows = []
            for row_num in range(1, min(11, max_row + 1)):
                row_data = []
                for col_num in range(1, min(6, max_col + 1)):
                    cell = ws.cell(row=row_num, column=col_num)
                    row_data.append({
                        'value': str(cell.value)[:50] if cell.value else None,
                        'type': str(type(cell.value).__name__),
                        'has_formula': cell.data_type == 'f'
                    })
                sample_rows.append({
                    'row': row_num,
                    'cells': row_data
                })
            
            # Sample last 10 rows
            sample_last_rows = []
            for row_num in range(max(1, max_row - 9), max_row + 1):
                row_data = []
                for col_num in range(1, min(3, max_col + 1)):
                    cell = ws.cell(row=row_num, column=col_num)
                    row_data.append({
                        'value': str(cell.value)[:50] if cell.value else None,
                        'type': str(type(cell.value).__name__)
                    })
                sample_last_rows.append({
                    'row': row_num,
                    'cells': row_data
                })
            
            # Find where actual data starts (look for dates)
            data_start_row = None
            for row_num in range(1, min(50, max_row + 1)):
                cell_value = ws.cell(row=row_num, column=1).value
                if cell_value and isinstance(cell_value, datetime):
                    data_start_row = row_num
                    break
                # Try parsing as date string
                try:
                    if cell_value and pd.to_datetime(cell_value, errors='coerce') is not pd.NaT:
                        data_start_row = row_num
                        break
                except:
                    pass
            
            inspection['sheet_details'].append({
                'sheet': sheet_name,
                'status': 'FOUND',
                'max_row': max_row,
                'max_column': max_col,
                'data_start_row': data_start_row,
                'first_10_rows': sample_rows,
                'last_10_rows': sample_last_rows
            })
        
        wb.close()
        os.remove(tmp_path)
        
        return jsonify(inspection)
        
    except Exception as e:
        logger.error(f"Error inspecting file: {e}")
        return jsonify({
            'error': str(e),
            'status': 'failed'
        }), 500

@app.route('/macro/backfill-yields-v2', methods=['POST'])
def backfill_yields_v2():
    """
    Smart backfill for Benchmark Yields Data sheet
    Handles descending date order and inserts at top
    """
    try:
        import openpyxl
        import tempfile
        from services.fred_api import FREDClient
        from datetime import timedelta
        
        logger.info("Starting Benchmark Yields backfill v2...")
        
        file_id = '1I3f36ghjh-NpI_EyhlZ9JTNUnGIWDkg4'
        
        # FRED series mapping to columns
        series_to_columns = {
            'DFF': 2,      # B - Fed Funds
            'DGS3MO': 4,   # D - 3mo
            'DGS6MO': 5,   # E - 6mo
            'DGS1': 6,     # F - 1yr
            'DGS2': 7,     # G - 2yr
            'DGS3': 8,     # H - 3yr
            'DGS5': 9,     # I - 5yr
            'DGS7': 10,    # J - 7yr
            'DGS10': 11,   # K - 10yr
            'DGS20': 12,   # L - 20yr
            'DGS30': 13,   # M - 30yr
        }
        
        # Initialize FRED client
        client = FREDClient()
        
        # Download file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        logger.info(f"Downloading file {file_id}...")
        google_drive.download_file(file_id, tmp_path)
        
        # Load workbook
        wb = openpyxl.load_workbook(tmp_path)
        
        if 'Data' not in wb.sheetnames:
            wb.close()
            os.remove(tmp_path)
            return jsonify({'error': 'Data sheet not found'}), 404
        
        ws = wb['Data']
        
        # Check current last date (should be in A4)
        current_last_date = ws.cell(row=4, column=1).value
        logger.info(f"Current last date in file: {current_last_date}")
        
        if not current_last_date:
            wb.close()
            os.remove(tmp_path)
            return jsonify({'error': 'No date found in A4'}), 400
        
        # Fetch data from FRED for all series
        logger.info("Fetching data from FRED...")
        
        # We need data from day after current_last_date to today
        start_date_dt = current_last_date + timedelta(days=1)
        end_date_dt = datetime.now()
        
        logger.info(f"Fetching data from {start_date_dt.strftime('%Y-%m-%d')} to {end_date_dt.strftime('%Y-%m-%d')}")
        
        # Fetch each series
        all_series_data = {}
        for series_id in series_to_columns.keys():
            try:
                # CRITICAL: Use DESC to get most recent data!
                data = client.get_series(series_id, limit=2000, sort_order='desc')
                
                if data and 'observations' in data:
                    df = pd.DataFrame(data['observations'])
                    
                    # Convert to datetime and normalize (remove timezone info)
                    df['date'] = pd.to_datetime(df['date']).dt.tz_localize(None)
                    df['value'] = pd.to_numeric(df['value'], errors='coerce')
                    
                    total_before_filter = len(df)
                    
                    # Log first and last dates in raw data
                    if len(df) > 0:
                        logger.info(f"{series_id} raw data range: {df['date'].min()} to {df['date'].max()}")
                        logger.info(f"  Looking for dates >= {start_date_dt} and <= {end_date_dt}")
                    
                    # Filter to our date range - using datetime objects
                    df = df[df['date'] >= start_date_dt]
                    after_start_filter = len(df)
                    
                    df = df[df['date'] <= end_date_dt]
                    after_end_filter = len(df)
                    
                    df = df.dropna(subset=['value'])
                    after_dropna = len(df)
                    
                    all_series_data[series_id] = df
                    logger.info(f"{series_id}: total={total_before_filter}, after_start={after_start_filter}, after_end={after_end_filter}, after_dropna={after_dropna}")
                    logger.info(f"Fetched {len(df)} observations for {series_id}")
                    
                    if len(df) > 0:
                        logger.info(f"  Date range in data: {df['date'].min()} to {df['date'].max()}")
                else:
                    logger.warning(f"No data for {series_id}")
                    
            except Exception as e:
                logger.error(f"Error fetching {series_id}: {e}")
        
        # Get union of all dates (business days where we have data)
        all_dates = set()
        for df in all_series_data.values():
            all_dates.update(df['date'].tolist())
        
        # Sort dates in DESCENDING order (newest first)
        sorted_dates = sorted(list(all_dates), reverse=True)
        
        logger.info(f"Total unique dates to add: {len(sorted_dates)}")
        
        if len(sorted_dates) == 0:
            wb.close()
            os.remove(tmp_path)
            return jsonify({
                'status': 'success',
                'message': 'No new data to add - file is up to date',
                'rows_added': 0
            })
        
        # Insert rows at row 4 (pushing existing data down)
        logger.info(f"Inserting {len(sorted_dates)} rows at row 4...")
        ws.insert_rows(4, len(sorted_dates))
        
        # Fill in the data
        rows_added = 0
        for i, date in enumerate(sorted_dates):
            row_num = 4 + i
            
            # Set date in column A
            ws.cell(row=row_num, column=1, value=date)
            
            # Set values for each series
            for series_id, col_num in series_to_columns.items():
                if series_id in all_series_data:
                    series_df = all_series_data[series_id]
                    # Find value for this date
                    value_rows = series_df[series_df['date'] == date]
                    if len(value_rows) > 0:
                        value = value_rows.iloc[0]['value']
                        ws.cell(row=row_num, column=col_num, value=value)
            
            rows_added += 1
            
            if rows_added % 100 == 0:
                logger.info(f"Progress: {rows_added}/{len(sorted_dates)} rows")
        
        logger.info(f"Filled {rows_added} rows with data")
        
        # Save workbook
        logger.info("Saving workbook...")
        wb.save(tmp_path)
        wb.close()
        
        # Upload back to Drive
        logger.info("Uploading to Google Drive...")
        google_drive.upload_file(tmp_path, file_id=file_id)
        
        # Clean up
        os.remove(tmp_path)
        
        return jsonify({
            'status': 'success',
            'message': f'Successfully backfilled {rows_added} business days',
            'file_id': file_id,
            'date_range': f"{sorted_dates[-1].strftime('%Y-%m-%d')} to {sorted_dates[0].strftime('%Y-%m-%d')}",
            'rows_added': rows_added,
            'series_updated': list(series_to_columns.keys()),
            'start_date': start_date_dt.strftime('%Y-%m-%d'),
            'end_date': end_date_dt.strftime('%Y-%m-%d'),
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error in yields backfill v2: {e}")
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc(),
            'status': 'failed'
        }), 500

@app.route('/macro/backfill-yields-correct', methods=['POST'])
def backfill_yields_correct():
    """
    CORRECT backfill - OVERWRITES rows (no inserting = no blank rows!)
    Includes TIPS data
    Preserves formulas
    """
    try:
        import openpyxl
        import tempfile
        from services.fred_api import FREDClient
        
        logger.info("Starting CORRECT Benchmark Yields backfill...")
        
        file_id = '1I3f36ghjh-NpI_EyhlZ9JTNUnGIWDkg4'
        
        # ALL series including TIPS AND 1-MONTH!
        series_to_columns = {
            'DFF': 2,         # B - Fed Funds
            'DGS1MO': 3,      # C - 1mo ← ADDED!
            'DGS3MO': 4,      # D - 3mo
            'DGS6MO': 5,      # E - 6mo
            'DGS1': 6,        # F - 1yr
            'DGS2': 7,        # G - 2yr
            'DGS3': 8,        # H - 3yr
            'DGS5': 9,        # I - 5yr
            'DGS7': 10,       # J - 7yr
            'DGS10': 11,      # K - 10yr
            'DGS20': 12,      # L - 20yr
            'DGS30': 13,      # M - 30yr
            'DFII5': 14,      # N - 5yr TIPS
            'DFII7': 15,      # O - 7yr TIPS
            'DFII10': 16,     # P - 10yr TIPS
            'DFII20': 17,     # Q - 20yr TIPS
            'DFII30': 18,     # R - 30yr TIPS
        }
        
        client = FREDClient()
        
        # Download file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        logger.info(f"Downloading file {file_id}...")
        google_drive.download_file(file_id, tmp_path)
        
        # Load workbook
        wb = openpyxl.load_workbook(tmp_path)
        
        if 'Data' not in wb.sheetnames:
            wb.close()
            os.remove(tmp_path)
            return jsonify({'error': 'Data sheet not found'}), 404
        
        ws = wb['Data']
        
        # Fetch ALL recent data (last 2000 = ~8 years)
        logger.info("Fetching data from FRED...")
        
        all_series_data = {}
        for series_id in series_to_columns.keys():
            try:
                data = client.get_series(series_id, limit=2000, sort_order='desc')
                
                if data and 'observations' in data:
                    df = pd.DataFrame(data['observations'])
                    df['date'] = pd.to_datetime(df['date']).dt.tz_localize(None)
                    df['value'] = pd.to_numeric(df['value'], errors='coerce')
                    df = df.dropna(subset=['date', 'value'])
                    
                    all_series_data[series_id] = df
                    logger.info(f"{series_id}: {len(df)} obs from {df['date'].min().date()} to {df['date'].max().date()}")
                    
            except Exception as e:
                logger.error(f"Error fetching {series_id}: {e}")
        
        # Get all unique dates
        all_dates = set()
        for df in all_series_data.values():
            all_dates.update(df['date'].tolist())
        
        # Sort DESCENDING (newest first)
        sorted_dates = sorted(list(all_dates), reverse=True)
        
        logger.info(f"Total dates: {len(sorted_dates)} from {sorted_dates[-1].date()} to {sorted_dates[0].date()}")
        
        # OVERWRITE starting at row 4 (NO INSERTING!)
        logger.info("Writing data (OVERWRITE mode - no blank rows)...")
        
        for i, date in enumerate(sorted_dates):
            row_num = 4 + i
            
            ws.cell(row=row_num, column=1, value=date)
            
            for series_id, col_num in series_to_columns.items():
                if series_id in all_series_data:
                    series_df = all_series_data[series_id]
                    matches = series_df[series_df['date'] == date]
                    if len(matches) > 0:
                        ws.cell(row=row_num, column=col_num, value=float(matches.iloc[0]['value']))
            
            if (i + 1) % 200 == 0:
                logger.info(f"Progress: {i+1}/{len(sorted_dates)}")
        
        logger.info(f"Wrote {len(sorted_dates)} rows")
        
        # Save workbook
        logger.info("Saving workbook...")
        wb.save(tmp_path)
        wb.close()
        
        # Upload back to Drive
        logger.info("Uploading to Google Drive...")
        google_drive.upload_file(tmp_path, file_id=file_id)
        
        # Clean up
        os.remove(tmp_path)
        
        return jsonify({
            'status': 'success',
            'message': f'Overwrote {len(sorted_dates)} rows with continuous data',
            'rows_written': len(sorted_dates),
            'date_range': f"{sorted_dates[-1].strftime('%Y-%m-%d')} to {sorted_dates[0].strftime('%Y-%m-%d')}",
            'series_count': len(series_to_columns),
            'includes_tips': True,
            'method': 'OVERWRITE (no inserting, no blank rows)',
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error in correct backfill: {e}")
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc(),
            'status': 'failed'
        }), 500

@app.route('/macro/backfill-yields-final', methods=['POST'])
def backfill_yields_final():
    """
    FINAL VERSION:
    - Only writes dates with complete main Treasury data (2yr, 5yr, 10yr, 30yr)
    - Creates professional charts from scratch
    - No blank rows in main series
    """
    try:
        import openpyxl
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.marker import Marker
        import tempfile
        from services.fred_api import FREDClient
        
        logger.info("Starting FINAL Benchmark Yields backfill with chart creation...")
        
        file_id = '1I3f36ghjh-NpI_EyhlZ9JTNUnGIWDkg4'
        
        # Define MAIN series (must have complete data for a date to be included)
        main_series = ['DGS2', 'DGS5', 'DGS10', 'DGS30']
        
        # All series with column mappings
        all_series = {
            'DFF': 2,         # B - Fed Funds
            'DGS1MO': 3,      # C - 1mo (supplementary)
            'DGS3MO': 4,      # D - 3mo
            'DGS6MO': 5,      # E - 6mo
            'DGS1': 6,        # F - 1yr
            'DGS2': 7,        # G - 2yr (MAIN)
            'DGS3': 8,        # H - 3yr
            'DGS5': 9,        # I - 5yr (MAIN)
            'DGS7': 10,       # J - 7yr
            'DGS10': 11,      # K - 10yr (MAIN)
            'DGS20': 12,      # L - 20yr
            'DGS30': 13,      # M - 30yr (MAIN)
            'DFII5': 14,      # N - 5yr TIPS (supplementary)
            'DFII7': 15,      # O - 7yr TIPS (supplementary)
            'DFII10': 16,     # P - 10yr TIPS (supplementary)
            'DFII20': 17,     # Q - 20yr TIPS (supplementary)
            'DFII30': 18,     # R - 30yr TIPS (supplementary)
        }
        
        client = FREDClient()
        
        # Download file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        logger.info(f"Downloading file...")
        google_drive.download_file(file_id, tmp_path)
        
        wb = openpyxl.load_workbook(tmp_path)
        
        if 'Data' not in wb.sheetnames:
            wb.close()
            os.remove(tmp_path)
            return jsonify({'error': 'Data sheet not found'}), 404
        
        ws = wb['Data']
        
        # Fetch data from FRED
        logger.info("Fetching data from FRED...")
        
        all_series_data = {}
        for series_id in all_series.keys():
            try:
                data = client.get_series(series_id, limit=2000, sort_order='desc')
                
                if data and 'observations' in data:
                    df = pd.DataFrame(data['observations'])
                    df['date'] = pd.to_datetime(df['date']).dt.tz_localize(None)
                    df['value'] = pd.to_numeric(df['value'], errors='coerce')
                    df = df.dropna(subset=['date', 'value'])
                    
                    all_series_data[series_id] = df
                    logger.info(f"{series_id}: {len(df)} obs")
                    
            except Exception as e:
                logger.error(f"Error fetching {series_id}: {e}")
        
        # Get intersection of dates for MAIN series only
        main_dates = None
        for series_id in main_series:
            if series_id in all_series_data:
                series_dates = set(all_series_data[series_id]['date'].tolist())
                if main_dates is None:
                    main_dates = series_dates
                else:
                    main_dates = main_dates.intersection(series_dates)
        
        if not main_dates:
            wb.close()
            os.remove(tmp_path)
            return jsonify({'error': 'No complete data for main series'}), 500
        
        # Sort DESCENDING
        sorted_dates = sorted(list(main_dates), reverse=True)
        
        logger.info(f"Writing {len(sorted_dates)} complete rows...")
        
        # Write data
        for i, date in enumerate(sorted_dates):
            row_num = 4 + i
            
            ws.cell(row=row_num, column=1, value=date)
            
            for series_id, col_num in all_series.items():
                if series_id in all_series_data:
                    series_df = all_series_data[series_id]
                    matches = series_df[series_df['date'] == date]
                    if len(matches) > 0:
                        ws.cell(row=row_num, column=col_num, value=float(matches.iloc[0]['value']))
            
            if (i + 1) % 500 == 0:
                logger.info(f"Progress: {i+1}/{len(sorted_dates)}")
        
        logger.info(f"Wrote {len(sorted_dates)} rows")
        
        # Create professional charts
        logger.info("Creating charts...")
        
        # Find or create chart sheets
        chart_configs = [
            {
                'sheet_name': '10yr',
                'title': '10-Year Treasury Yield',
                'data_col': 11,  # Column K
                'color': '0070C0',
                'y_axis_title': 'Yield (%)'
            },
            {
                'sheet_name': '2yr',
                'title': '2-Year Treasury Yield', 
                'data_col': 7,  # Column G
                'color': 'C00000',
                'y_axis_title': 'Yield (%)'
            },
            {
                'sheet_name': '5yr',
                'title': '5-Year Treasury Yield',
                'data_col': 9,  # Column I
                'color': '00B050',
                'y_axis_title': 'Yield (%)'
            },
            {
                'sheet_name': '30yr',
                'title': '30-Year Treasury Yield',
                'data_col': 13,  # Column M
                'color': 'FFC000',
                'y_axis_title': 'Yield (%)'
            },
        ]
        
        for config in chart_configs:
            sheet_name = config['sheet_name']
            
            # Create or get sheet
            if sheet_name in wb.sheetnames:
                chart_ws = wb[sheet_name]
                # Clear existing charts
                chart_ws._charts = []
            else:
                chart_ws = wb.create_sheet(sheet_name)
            
            # Create chart
            chart = LineChart()
            chart.title = config['title']
            chart.style = 2
            chart.y_axis.title = config['y_axis_title']
            chart.x_axis.title = 'Date'
            
            # Data references (from Data sheet)
            data = Reference(ws, min_col=config['data_col'], min_row=4, max_row=4+len(sorted_dates)-1)
            dates = Reference(ws, min_col=1, min_row=4, max_row=4+len(sorted_dates)-1)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(dates)
            
            # Styling
            series = chart.series[0]
            series.graphicalProperties.line.solidFill = config['color']
            series.graphicalProperties.line.width = 20000  # Line width
            series.smooth = True  # Smooth line
            
            # Remove markers for cleaner look
            series.marker = Marker('none')
            
            # Chart size
            chart.width = 20
            chart.height = 10
            
            # Add to sheet
            chart_ws.add_chart(chart, "A1")
            
            logger.info(f"Created chart for {sheet_name}")
        
        # Create Yield Curve chart
        logger.info("Creating Yield Curve chart...")
        
        if 'Yield Curve' in wb.sheetnames:
            curve_ws = wb['Yield Curve']
            curve_ws._charts = []
        else:
            curve_ws = wb.create_sheet('Yield Curve')
        
        curve_chart = LineChart()
        curve_chart.title = 'Treasury Yield Curve (Current)'
        curve_chart.style = 2
        curve_chart.y_axis.title = 'Yield (%)'
        curve_chart.x_axis.title = 'Maturity'
        
        # Use most recent date (row 4) for all maturities
        # Create data on the curve sheet itself
        maturities = ['3mo', '6mo', '1yr', '2yr', '3yr', '5yr', '7yr', '10yr', '20yr', '30yr']
        columns = [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]  # D through M
        
        # Write maturity labels and values
        for idx, (mat, col) in enumerate(zip(maturities, columns)):
            curve_ws.cell(row=1, column=idx+1, value=mat)
            # Get value from Data sheet row 4
            value = ws.cell(row=4, column=col).value
            curve_ws.cell(row=2, column=idx+1, value=value)
        
        # Create chart from this data
        data = Reference(curve_ws, min_col=1, min_row=2, max_col=len(maturities))
        cats = Reference(curve_ws, min_col=1, min_row=1, max_col=len(maturities))
        
        curve_chart.add_data(data, titles_from_data=False)
        curve_chart.set_categories(cats)
        
        series = curve_chart.series[0]
        series.graphicalProperties.line.solidFill = '0070C0'
        series.graphicalProperties.line.width = 25000
        series.smooth = False
        
        curve_chart.width = 20
        curve_chart.height = 10
        
        curve_ws.add_chart(curve_chart, "A4")
        
        logger.info("Created Yield Curve chart")
        
        # Save
        logger.info("Saving workbook...")
        wb.save(tmp_path)
        wb.close()
        
        # Upload
        logger.info("Uploading to Google Drive...")
        google_drive.upload_file(tmp_path, file_id=file_id)
        
        os.remove(tmp_path)
        
        return jsonify({
            'status': 'success',
            'message': f'Backfilled {len(sorted_dates)} complete rows + created charts',
            'rows_written': len(sorted_dates),
            'date_range': f"{sorted_dates[-1].strftime('%Y-%m-%d')} to {sorted_dates[0].strftime('%Y-%m-%d')}",
            'main_series': main_series,
            'strategy': 'Complete main series data only (no blank rows)',
            'charts_created': len(chart_configs) + 1,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error: {e}")
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/macro/analyze-data-sheet', methods=['GET'])
def analyze_data_sheet():
    """
    Deep analysis of the Data sheet to understand structure for backfilling
    """
    try:
        import openpyxl
        import tempfile
        
        file_id = '1I3f36ghjh-NpI_EyhlZ9JTNUnGIWDkg4'
        
        # Download file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        google_drive.download_file(file_id, tmp_path)
        
        # Load workbook
        wb = openpyxl.load_workbook(tmp_path, data_only=False)
        
        # Read Data sheet
        if 'Data' not in wb.sheetnames:
            wb.close()
            os.remove(tmp_path)
            return jsonify({'error': 'Data sheet not found', 'available_sheets': wb.sheetnames}), 404
        
        ws = wb['Data']
        
        analysis = {
            'sheet_name': 'Data',
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'structure': {}
        }
        
        # Read first 20 rows to understand structure
        first_rows = []
        for row_num in range(1, min(21, ws.max_row + 1)):
            row_data = {}
            for col_num in range(1, min(30, ws.max_column + 1)):
                cell = ws.cell(row=row_num, column=col_num)
                col_letter = openpyxl.utils.get_column_letter(col_num)
                
                cell_info = {
                    'value': str(cell.value)[:100] if cell.value else None,
                    'type': str(type(cell.value).__name__)
                }
                
                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                    cell_info['has_formula'] = True
                
                row_data[col_letter] = cell_info
            
            first_rows.append({
                'row': row_num,
                'data': row_data
            })
        
        analysis['first_20_rows'] = first_rows
        
        # Find where date column is
        date_col = None
        for col_num in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=10, column=col_num).value
            if cell_value and isinstance(cell_value, datetime):
                date_col = openpyxl.utils.get_column_letter(col_num)
                break
        
        if date_col:
            date_samples = []
            sample_rows = [10, 50, 100, 200, ws.max_row - 10, ws.max_row]
            for row_num in sample_rows:
                if row_num > 0 and row_num <= ws.max_row:
                    date_val = ws.cell(row=row_num, column=openpyxl.utils.column_index_from_string(date_col)).value
                    date_samples.append({
                        'row': row_num,
                        'date': str(date_val) if date_val else None
                    })
            
            analysis['date_column'] = {
                'column': date_col,
                'samples': date_samples
            }
        
        # Check for yield columns in row 2
        header_row = 2
        yield_columns = {}
        for col_num in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col_num).value
            if header and any(term in str(header).lower() for term in ['yr', 'mo', 'fed', 'tips', 'date']):
                col_letter = openpyxl.utils.get_column_letter(col_num)
                yield_columns[col_letter] = str(header)
        
        analysis['yield_columns'] = yield_columns
        
        wb.close()
        os.remove(tmp_path)
        
        return jsonify(analysis)
        
    except Exception as e:
        logger.error(f"Error analyzing Data sheet: {e}")
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

#═══════════════════════════════════════════════════════════════════════════════
# STOCK SCREENING ENDPOINTS (Week 3-5 implementation)
#═══════════════════════════════════════════════════════════════════════════════

@app.route('/stocks/test-screen', methods=['POST'])
def test_stock_screen():
    """
    Test endpoint - Returns dummy stock screening results
    Week 3-5: Replace with real screening from US_Sector_Data.xlsm
    """
    logger.info("Screening stocks (test mode - Week 1)")
    
    data = request.get_json() or {}
    sectors = data.get('sectors', ['Financials', 'Industrials'])
    
    # Dummy stock data
    all_stocks = [
        {
            'ticker': 'JPM',
            'company': 'JPMorgan Chase',
            'sector': 'Financials',
            'price': 156.50,
            'pe': 11.2,
            'roe': 17.0,
            'eps_growth_y1': 18.0,
            'eps_growth_y2': 24.0,
            'market_cap': 445000000000,
            'beta': 1.19,
            'eg_profile': 'Profile 1 - Accelerating Outperformer'
        },
        {
            'ticker': 'BAC',
            'company': 'Bank of America',
            'sector': 'Financials',
            'price': 32.45,
            'pe': 10.8,
            'roe': 14.0,
            'eps_growth_y1': 15.0,
            'eps_growth_y2': 22.0,
            'market_cap': 265000000000,
            'beta': 1.31,
            'eg_profile': 'Profile 1 - Accelerating Outperformer'
        },
        {
            'ticker': 'CAT',
            'company': 'Caterpillar',
            'sector': 'Industrials',
            'price': 214.50,
            'pe': 14.2,
            'roe': 21.0,
            'eps_growth_y1': 12.0,
            'eps_growth_y2': 18.0,
            'market_cap': 108000000000,
            'beta': 1.21,
            'eg_profile': 'Profile 2 - Stable Outperformer'
        }
    ]
    
    filtered = [s for s in all_stocks if s['sector'] in sectors]
    
    return jsonify({
        'candidates': filtered,
        'count': len(filtered),
        'sectors_requested': sectors,
        'status': 'success',
        'mode': 'test',
        'note': 'Week 1 - Dummy data. Week 3-5 will implement real screening.'
    })

#═══════════════════════════════════════════════════════════════════════════════
# UTILITY ENDPOINTS
#═══════════════════════════════════════════════════════════════════════════════

@app.route('/ping')
def ping():
    """Simple ping for monitoring"""
    return jsonify({
        'status': 'ok',
        'timestamp': datetime.now().isoformat()
    })

@app.route('/test/telegram', methods=['POST'])
def test_telegram():
    """Test Telegram notification"""
    try:
        data = request.get_json() or {}
        message = data.get('message', 'Test message from Trading System')
        
        # Import telegram handler
        from services import telegram_handler
        
        result = telegram_handler.send_message(message)
        
        return jsonify({
            'status': 'success',
            'message': 'Telegram message sent',
            'result': result
        })
        
    except Exception as e:
        logger.error(f"Telegram test failed: {e}")
        return jsonify({'error': str(e)}), 500

#═══════════════════════════════════════════════════════════════════════════════
# ERROR HANDLERS
#═══════════════════════════════════════════════════════════════════════════════

@app.errorhandler(404)
def not_found(error):
    return jsonify({
        'error': 'Endpoint not found',
        'available_endpoints': [
            'GET /',
            'GET /health',
            'GET /drive/list/<folder_type>',
            'GET /drive/read/<file_id>',
            'POST /templates/test-update',
            'GET /macro/templates',
            'GET /macro/test-fetch',
            'GET /macro/fetch',
            'GET /macro/audit-templates',
            'GET /macro/audit-templates-v2',
            'POST /macro/backfill-yields',
            'POST /macro/backfill-yields-v2',
            'POST /macro/backfill-yields-v3',
            'POST /macro/backfill-yields-correct',
            'POST /macro/backfill-yields-final',
            'GET /macro/inspect-yields',
            'GET /macro/analyze-data-sheet',
            'POST /stocks/test-screen',
            'POST /test/telegram'
        ]
    }), 404

@app.errorhandler(500)
def internal_error(error):
    logger.error(f"Internal server error: {error}")
    return jsonify({
        'error': 'Internal server error',
        'message': str(error)
    }), 500

#═══════════════════════════════════════════════════════════════════════════════
# MAIN
#═══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    
    logger.info(f"Starting Trading System API on port {port}")
    logger.info(f"Debug mode: {debug}")
    
    app.run(host='0.0.0.0', port=port, debug=debug)
